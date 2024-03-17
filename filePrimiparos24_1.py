################# kramer ###########
import pandas as pd
from openpyxl import load_workbook

rutaPrimiparo = "C:/Users/jodam/Downloads/POLITICA_GRATUIDAD/PIAM2024_1/INSUMODARCA/CI/PRIMIPAROSAENVIAR.xlsx"
primiparoConsolidado = 'TODOS'

def estadoLiquidacionPolitica(valor):
    if valor == 0:
        return 'Liquidacion sin politica'
    elif valor < 0:
        return 'Liquidacion con politica'
    elif pd.isnull(valor):
        return 'Sin liquidacion financiera'
    else:
        return 'Otro estado financiero'
    
def prevalidacionIESPolitica(estrato, sisben):
    if estrato in [1,2,3] and sisben in ["A", "B", "C"]:
        return 'Cumple estrato y sisben'
    elif estrato in [1,2,3]:
        return 'Cumple estrato'
    elif sisben in ["A", "B", "C"]:
        return 'Cumple sisben'
    else:
        return 'No cumple requisitos'
    
def calculadoradeMatricula(ruta,hoja):
    #dataframe = lectura(ruta,hoja)
    dataframe = pd.read_excel(ruta,sheet_name=hoja)
    dataframe['BRUTA']= dataframe['DERECHOS_MATRICULA'] + dataframe['BIBLIOTECA_DEPORTES'] +\
                        dataframe['LABORATORIOS'] + dataframe['RECURSOS_COMPUTACIONALES'] +\
                        dataframe['SEGURO_ESTUDIANTIL'] + dataframe['VRES_COMPLEMENTARIOS'] +\
                        dataframe['RESIDENCIAS'] +dataframe['REPETICIONES']
    dataframe['BRUTAORD']=  dataframe['DERECHOS_MATRICULA'] + dataframe['BIBLIOTECA_DEPORTES'] +\
                            dataframe['LABORATORIOS'] + dataframe['RECURSOS_COMPUTACIONALES'] +\
                            dataframe['VRES_COMPLEMENTARIOS'] + dataframe['RESIDENCIAS'] +\
                            dataframe['REPETICIONES']
    dataframe['MERITO']= -dataframe['CONVENIO_DESCENTRALIZACION'] - dataframe['BECA'] -\
                          dataframe['MATRICULA_HONOR'] - dataframe['MEDIA_MATRICULA_HONOR'] -\
                          dataframe['TRABAJO_GRADO'] - dataframe['DOS_PROGRAMAS'] -\
                          dataframe['DESCUENTO_HERMANO'] - dataframe['ESTIMULO_EMP_DTE_PLANTA'] -\
                          dataframe['EXEN_HIJOS_CONYUGE_CATEDRA'] - dataframe['EXEN_HIJOS_CONYUGE_OCASIONAL'] -\
                          dataframe['HIJOS_TRABAJADORES_OFICIALES'] - dataframe['ACTIVIDAES_LUDICAS_DEPOR'] -\
                          dataframe['DESCUENTOS'] - dataframe['SERVICIOS_RELIQUIDACION']
    dataframe['NETAORD']= dataframe['BRUTAORD'] + dataframe['VOTO']
    dataframe['NETA']= dataframe['BRUTA'] + dataframe['VOTO'] - dataframe['MERITO']
    dataframe['ESTADOLIQUIDACION'] = dataframe['GRATUIDAD_MATRICULA'].apply(estadoLiquidacionPolitica)
    dataframe['ESTADOPREVALIES'] = dataframe.apply(lambda row: prevalidacionIESPolitica(row['ESTRATO'], row['GRUPOSISBEN']), axis=1)
    
    with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=hoja, index=False)


"""def calculadoradeEstados(ruta,hoja):
    #dataframe = lectura(ruta,hoja)
    dataframe = pd.read_excel(ruta,sheet_name=hoja)
    dataframe['ESTADOPREVALIES'] = dataframe.apply(lambda row: prevalidacionIESPolitica(row['ESTRATO'], row['GRUPOSISBEN']), axis=1)
    #dataframe.to_excel(ruta, sheet_name=hoja, index=False)    
    with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        dataframe.to_excel(writer, sheet_name=hoja, index=False)
        
def copiarHojas(origen, destino):
    wb_origen = load_workbook(origen)
    wb_destino = load_workbook(destino)
    for sheet in wb_origen.sheetnames:
        ws = wb_origen[sheet]
        wb_destino.create_sheet(title=sheet)
        for row in ws.iter_rows():
            wb_destino[sheet].append([cell.value for cell in row])
    wb_destino.save(destino)"""

def lectura(ruta):
    
    xls = pd.ExcelFile(ruta, engine='openpyxl')
    hojas_disponibles = xls.sheet_names
        
    for hoja in hojas_disponibles:
        dataframe = pd.read_excel(ruta,sheet_name=hoja)
        
        if hoja == 'TODOS':
            calculadoradeMatricula(ruta,hoja)
                       
        else:
            with pd.ExcelWriter(ruta, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                dataframe.to_excel(writer, sheet_name=hoja, index=False)

       
lectura(rutaPrimiparo)
#calculadoradeMatricula(rutaPrimiparo,primiparoConsolidado)
#calculadoradeEstados(rutaPrimiparo,primiparoConsolidado)
#copiarHojas(rutaPrimiparo,rutaPrimiparo)


#print(lectura(rutaPrimiparo,primiparoConsolidado))
#info(rutaPrimiparo,primiparoConsolidado)